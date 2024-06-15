import openpyxl
import random
import matplotlib.pyplot as plt
import datetime


class Stock:
    def __init__(self, name, high_prices, low_prices, close_prices):
        self.name = name
        self.high_prices = high_prices
        self.low_prices = low_prices
        self.close_prices = close_prices
        self.history = []
        self.current_day = 0
        self.price = None  # Initialize the price to None

    def update_price(self):
        if self.current_day < len(self.close_prices):
            # Randomly pick between high, low, and close prices
            self.price = random.choice([self.high_prices[self.current_day], self.low_prices[self.current_day],
                                        self.close_prices[self.current_day]])
            self.history.append(self.price)
            self.current_day += 1
        else:
            raise ValueError("No more historical data available for update.")

    def get_price(self):
        if self.price is None:
            raise ValueError("Price has not been updated yet.")
        return self.price

    def get_history(self):
        return self.history


class Portfolio:
    def __init__(self):
        self.stocks = {}
        self.cash = 1300  # initial cash in portfolio
        self.history = []

    def buy_stock(self, stock, amount):
        cost = stock.get_price() * amount
        if self.cash >= cost:
            self.cash -= cost
            if stock.name in self.stocks:
                self.stocks[stock.name]['amount'] += amount
            else:
                self.stocks[stock.name] = {'stock': stock, 'amount': amount}
            print(f"Bought {amount} shares of {stock.name} at {stock.get_price():.2f} each, total cost: {cost:.2f}")
        else:
            print("Not enough cash to buy stock.")

    def sell_stock(self, stock_name, amount):
        if stock_name in self.stocks and self.stocks[stock_name]['amount'] >= amount:
            stock = self.stocks[stock_name]['stock']
            revenue = stock.get_price() * amount
            self.cash += revenue
            self.stocks[stock_name]['amount'] -= amount
            if self.stocks[stock_name]['amount'] == 0:
                del self.stocks[stock_name]
            print(f"Sold {amount} shares of {stock_name} at {stock.get_price():.2f} each, total revenue: {revenue:.2f}")
        else:
            print("Not enough shares to sell or stock not in portfolio.")

    def update_portfolio_value(self):
        total_value = self.cash
        for stock_data in self.stocks.values():
            total_value += stock_data['stock'].get_price() * stock_data['amount']
        self.history.append(total_value)
        return total_value

    def get_history(self):
        return self.history


# Load Excel workbook and worksheet
# C:\\Users\\kingt\\OneDrive\\Desktop\\New folder\\NVDA.xlsx for Frank's location
# /home/jera1470/PycharmProjects/Python-trading-algo/NVDA.xlsx for Jera's location
wb = openpyxl.load_workbook("/home/jera1470/PycharmProjects/Python-trading-algo/NVDA.xlsx")
ws = wb.active


# Function to ensure data is numerical
def get_numerical_data(cell_value):
    if isinstance(cell_value, (int, float)):
        return cell_value
    elif isinstance(cell_value, datetime.datetime):
        return cell_value.timestamp()  # Convert datetime to a numerical timestamp if necessary
    elif isinstance(cell_value, str):
        try:
            return float(cell_value.replace(',', ''))
        # Gives error if it cannot find a numeric value
        except ValueError:
            raise ValueError(f"Non-numeric data found in the Excel file: {cell_value}")
    else:
        raise ValueError(f"Non-numeric data found in the Excel file: {cell_value}")


# Historical data for 30 days: high, low, and close prices

# TODO: null values to have exceptions to data (in get_numerical_data)
high_prices_a = [get_numerical_data(ws.cell(row=i+1, column=2).value) for i in range(1, 251)]
low_prices_a = [get_numerical_data(ws.cell(row=i+1, column=3).value) for i in range(1, 251)]
close_prices_a = [get_numerical_data(ws.cell(row=i+1, column=5).value) for i in range(1, 251)]

high_prices_b = [105, 106, 107, 108, 109, 110, 109, 108, 107, 106, 105, 104, 103, 102, 101, 100, 99, 98, 97, 96, 95, 94,
                 93, 92, 91, 90, 89, 88, 87, 86]
low_prices_b = [95, 96, 97, 98, 99, 100, 99, 98, 97, 96, 95, 94, 93, 92, 91, 90, 89, 88, 87, 86, 85, 84, 83, 82, 81, 80,
                79, 78, 77, 76]
close_prices_b = [100, 101, 102, 103, 104, 105, 104, 103, 102, 101, 100, 99, 98, 97, 96, 95, 94, 93, 92, 91, 90, 89, 88,
                  87, 86, 85, 84, 83, 82, 81]

high_prices_c = [205, 207, 209, 208, 210, 212, 214, 213, 212, 211, 210, 209, 208, 207, 206, 205, 204, 203, 202, 201,
                 200, 199, 198, 197, 196, 195, 194, 193, 192, 191]
low_prices_c = [195, 197, 199, 198, 200, 202, 204, 203, 202, 201, 200, 199, 198, 197, 196, 195, 194, 193, 192, 191, 190,
                189, 188, 187, 186, 185, 184, 183, 182, 181]
close_prices_c = [200, 202, 204, 203, 205, 207, 209, 208, 207, 206, 205, 204, 203, 202, 201, 200, 199, 198, 197, 196,
                  195, 194, 193, 192, 191, 190, 189, 188, 187, 186]

# Initialize stocks
stock_a = Stock("TechCorp", high_prices_a, low_prices_a, close_prices_a)
stock_b = Stock("HealthInc", high_prices_b, low_prices_b, close_prices_b)
stock_c = Stock("FinanceCo", high_prices_c, low_prices_c, close_prices_c)

# Initialize portfolio
portfolio = Portfolio()

# Simulate market and portfolio actions
for day in range(30):  # simulate for 30 days
    print(f"Day {day + 1}")

    # Update stock prices
    stock_a.update_price()
    stock_b.update_price()
    stock_c.update_price()

    # Buy or sell stocks randomly
    action = random.choice(['buy', 'sell', 'hold'])
    if action == 'buy':
        stock_choice = random.choice([stock_a, stock_b, stock_c])
        portfolio.buy_stock(stock_choice, random.randint(1, 5))
    elif action == 'sell':
        stock_name_choice = random.choice([stock_a.name, stock_b.name, stock_c.name])
        portfolio.sell_stock(stock_name_choice, random.randint(1, 5))

    # Update and print portfolio value
    current_value = portfolio.update_portfolio_value()
    print(f"Portfolio value: {current_value:.2f}\n")

# Plot portfolio value over time
plt.plot(portfolio.get_history())
plt.title('Portfolio Value Over Time')
plt.xlabel('Days')
plt.ylabel('Value')
plt.show()
